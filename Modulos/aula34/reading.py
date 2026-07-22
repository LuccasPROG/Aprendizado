# openpyxl - Ler e alterar dados de uma planilha
# Com essa biblioteca será possível ler e escrever dados em células
# específicas, formatar células, inserir gráficos,
# criar fórmulas, adicionar imagens e outros elementos gráficos às suas
# planilhas. Ela é útil para automatizar tarefas envolvendo planilhas do
# Excel, como a criação de relatórios e análise de dados e/ou facilitando a
# manipulação de grandes quantidades de informações.
# Instalação necessária: pip install openpyxl
# Documentação: https://openpyxl.readthedocs.io/en/stable/

from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.cell import Cell
from openpyxl.worksheet import worksheet

ROOT_FOLDER = Path(__file__).parent
WORKBOOK_PATH = ROOT_FOLDER / 'workbook.xlsx'

#Carregando um arquivo do execel
workbook:Workbook = load_workbook(WORKBOOK_PATH)
#Nome para a planilha
sheet_name = 'Minha Planilha'
#seleciona a planilha
worksheet: worksheet = workbook[sheet_name]  #type:ignore

row : tuple[Cell]
for row in worksheet.iter_rows():
    for col in row:
        print(col.value, end='\t')

        if col.value == 'Lucas':
            worksheet.cell(col.row, 2, 19)
    print()

# worksheet['B2'].value = 20

        


workbook.save(WORKBOOK_PATH) #save 