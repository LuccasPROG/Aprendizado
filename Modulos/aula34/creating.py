# openpyxl para arquivos Excel xlsx, xlsm, xltx e xltm (instalação)
# Com essa biblioteca será possível ler e escrever dados em células
# específicas, formatar células, inserir gráficos,
# criar fórmulas, adicionar imagens e outros elementos gráficos às suas
# planilhas. Ela é útil para automatizar tarefas envolvendo planilhas do
# Excel, como a criação de relatórios e análise de dados e/ou facilitando a
# manipulação de grandes quantidades de informações.
# Instalação necessária: pip install openpyxl
# Documentação: https://openpyxl.readthedocs.io/en/stable/
from pathlib import Path
from openpyxl import Workbook
from openpyxl.worksheet import worksheet

ROOT_FOLDER = Path(__file__).parent
WORKBOOK_PATH = ROOT_FOLDER / 'workbook.xlsx'

#Nome para a planilha
workbook = Workbook()
sheet_name = 'Minha Planilha'
#Cria a Planilha
workbook.create_sheet(sheet_name, 0)
#seleciona a planilha
worksheet: worksheet = workbook[sheet_name]  #type:ignore

#REMOVER Planilha
workbook.remove(workbook['Sheet'])
 
worksheet.cell(1, 1, 'Name')
worksheet.cell(1, 2, 'Age')
worksheet.cell(1, 3, 'notice')


students = [
    #name           age    notice
    ['Lucas',        19,     5.8],
    ['Valentina',    8,     4.8],
    ['Suelen',       35,     8.4],
    ['Edna',         18,     9.4],
    ['Edson',        50,     7.8],

]
#       'FORMA DIFICIL'
# for i, student_row in enumerate(students, start=2):
#     for j , students_column in enumerate(student_row, start=1):
#         worksheet.cell(i, j, students_column)

#FORMA FACIL
for student in students:
    worksheet.append(student)
        


workbook.save(WORKBOOK_PATH) #save 